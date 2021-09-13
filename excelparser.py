# Source Code is copyright of Anaana Terdoo George, aka. Infinime. C.2021, All Rights Reserved, Distributed under a closed source license.
with open("settings.txt") as settings:
    EssayFile, ExcelWordsFile = list(filter(("").__ne__, str(settings.read()).replace(
        'Essay File Path:', '').replace("Excel File Path:", "").split("\n")))
    EssayFile = EssayFile.strip()
    ExcelWordsFile = ExcelWordsFile.strip()
print("Reading From:", EssayFile,
      "\nComparing to the Word List at:", ExcelWordsFile, "\n\n")

import openpyxl
import codecs
import pdfkit
import matplotlib.pyplot as plt

EssayFile = EssayFile.replace("\\", "\\\\")
ExcelWordsFile = ExcelWordsFile.replace("\\", "\\\\")
book = openpyxl.load_workbook(filename=ExcelWordsFile)
with codecs.open(EssayFile, 'r', encoding="UTF-8", errors='ignore') as e:
    essay = str(e.read())
sheets = book.sheetnames
category1 = []
category2 = []
category3 = []
category4 = []
category5 = []
category6 = []
category7 = []
category8 = []
category9 = []
category10 = []
category11 = []
category12 = []
category13 = []
category14 = []
category15 = []
category16 = []
category17 = str(essay).split(" ")
deletethese = []
categories = [category1, category2, category3, category4,
              category5, category6, category7, category8,
              category9, category10, category11, category12,
              category13, category14, category15, category16]
# categoryStrings=['category1','category2','category3','category4','category5','category6','category7','category8','category9','category10','category11','category12','category13','category14','category15','category16']
numberInCategory = dict(
    zip(sheets, [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0]))
# print(len(sheets))
for sheetnum in range(16):
    currentSheet = book[sheets[sheetnum]]
    for row in range(1, currentSheet.max_row + 1):
        categories[sheetnum] += [str(currentSheet["B" +
                                                  str(row)].value).strip()]

for n in range(16):
    for y in categories[n]:
        for x in category17:
            if str(y).lower() in x.lower():
                numberInCategory[sheets[n]] += 1
                deletethese += [x]
            if '\\' in x or '' == x:
                deletethese += [x]

for x in deletethese:
    category17 = list(filter((x).__ne__, category17))

for x in range(16):
    print("The number of words in sheet '" +
          sheets[x] + "' is: ", numberInCategory[sheets[x]])
print("The number of words not in any sheets is:", len(category17), "\n\n")

plt.bar(['1', '2', '3', '4', '5', '6', '7', '8', '9', '10', '11', '12', '13',
         '14', "15", "16", "Nil"], list(numberInCategory.values()) + [len(category17)])
plt.title('Word Frequency Chart')
plt.xlabel('Word Category')
plt.ylabel('Category Frequency')
plt.savefig(str(EssayFile.split("\\")[-1:][0]) + " Word Frequency Graph.png", dpi=600)
print("Bar Chart generated successfully, saving...")
print("Bar chart saved at", str(EssayFile.split(
    "\\")[-1:][0]) + " Word Frequency Graph.png!\n")


def selectBold(text, level):
    arr = []
    if level < 16:
        for x in range(level, 16):
            arr += categories[x]
    arr += category17
    for y in arr:
        text = text.replace(" " + y + " ", "<b> " + y + " </b>")
    # print(text)
    return text


if input("Do you want to select the rarer words in this document? (Y/N)\n").upper() == "Y":
    level = str(input("Okay. Words above level? (1-16)\n"))
    # print(level)
    essay = selectBold(essay, int(level))
    essay = '<meta http-equiv="Content-type" content="text/html; charset=utf-8" />' + essay
    pdfkit.from_string(essay, str(EssayFile.split(
        "\\")[-1:][0]) + " Words above level " + level + ".pdf")
    print("PDF file generated successfully, formatting...")
    print("PDF saved at", str(EssayFile.split("\\")[-1:][0]) + " Words above level " + level + ".pdf!")
    input("Press <ENTER> to exit the program")
